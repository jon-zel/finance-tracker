"""
All filesystem/Excel access for trip files lives here — nowhere else.

Golden Rule (spec §2): the .xlsx files are a permanent contract.
- Read columns by header name, never by position.
- Never touch cells in columns/rows the app doesn't own.
- Rows are append-only from the app.

Trip storage is per-user (deployment-auth-spec §2A): every function below
takes the caller's sanitized folder name (see auth.email_to_folder) and all
of them resolve paths only under that user's own subfolder of TRIPS_DIR,
with a path-traversal check (§2A.3) enforced on every call.
"""
import glob
import logging
import os
import re
import threading
from datetime import date, datetime
from pathlib import Path

import openpyxl

import config

logger = logging.getLogger("travel_expenses.security")

REQUIRED_HEADERS = ["Date", "Amount", "Category", "Notes"]

TRIP_NAME_MAX_LEN = 100
_CONTROL_CHARS_RE = re.compile(r"[\x00-\x1f]")

# FastAPI runs these sync route handlers in a thread pool, so two requests
# (e.g. two browser tabs, or a rapid double-submit) can genuinely race here:
# both threads can load_workbook() before either calls save(), and whichever
# save() lands second silently discards the first thread's new row/trip —
# a smaller-scale case of the same "stale snapshot overwrites disk" data-loss
# class found in the finance-tracker app. One process-wide lock serializes all
# trip-file writes; this app is small-scale/family so throughput isn't a concern.
_write_lock = threading.Lock()


class TripNotFoundError(Exception):
    pass


class TripAlreadyExistsError(Exception):
    pass


class InvalidTripFileError(Exception):
    pass


class InvalidTripNameError(Exception):
    pass


def ensure_user_folder(user_folder: str) -> None:
    """Creates the user's trips folder if it doesn't exist yet (spec §2A.1,
    acceptance criteria #20). Called on every successful login."""
    os.makedirs(os.path.join(config.TRIPS_DIR, user_folder), exist_ok=True)


def validate_trip_name(name: str) -> None:
    """Path traversal defense, spec §2A.3 step 1. Applied to every trip name
    coming from the frontend before it touches the filesystem. Callers at
    the route layer should call this first (and log on rejection, spec
    §2A.3/§11 AC#19) — it never touches disk, so it's safe to call before
    any existence check."""
    if not name or len(name) > TRIP_NAME_MAX_LEN:
        raise InvalidTripNameError(name)
    if "/" in name or "\\" in name or ".." in name:
        raise InvalidTripNameError(name)
    if _CONTROL_CHARS_RE.search(name):
        raise InvalidTripNameError(name)
    if name.endswith(" ") or name.endswith("."):
        raise InvalidTripNameError(name)


def _trip_path(user_folder: str, name: str) -> str:
    """Builds and verifies a trip file path (spec §2A.3 steps 2-3). Re-runs
    validate_trip_name as defense-in-depth even though route handlers
    already call it — this function is the last line of defense before any
    actual filesystem access."""
    validate_trip_name(name)

    user_dir = Path(config.TRIPS_DIR) / user_folder
    built_path = user_dir / f"{name}.xlsx"

    resolved_user_dir = user_dir.resolve()
    resolved_built_path = built_path.resolve()
    if not str(resolved_built_path).startswith(str(resolved_user_dir) + os.sep):
        logger.warning("Path traversal attempt blocked: user_folder=%r name=%r", user_folder, name)
        raise InvalidTripNameError(name)

    return str(built_path)


def list_trips(user_folder: str) -> list[str]:
    user_dir = os.path.join(config.TRIPS_DIR, user_folder)
    os.makedirs(user_dir, exist_ok=True)
    paths = glob.glob(os.path.join(user_dir, "*.xlsx"))
    names = [os.path.splitext(os.path.basename(p))[0] for p in paths]
    return sorted(names, key=str.lower)


def trip_exists(user_folder: str, name: str) -> bool:
    try:
        path = _trip_path(user_folder, name)
    except InvalidTripNameError:
        return False
    return os.path.exists(path)


def create_trip(user_folder: str, name: str) -> None:
    path = _trip_path(user_folder, name)
    with _write_lock:
        if os.path.exists(path):
            raise TripAlreadyExistsError(name)
        os.makedirs(os.path.dirname(path), exist_ok=True)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(REQUIRED_HEADERS)
        wb.save(path)


def _header_column_map(ws) -> dict[str, int]:
    """Maps header name -> 1-indexed column number, read from row 1."""
    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None:
            headers[str(cell.value).strip()] = cell.column
    return headers


def _require_headers(headers: dict[str, int], trip_name: str) -> None:
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise InvalidTripFileError(
            f"'{trip_name}.xlsx' is missing required column(s): {', '.join(missing)}"
        )


def _cell_to_iso_date(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return ""
    # Tolerate manually-typed text dates in the user's file.
    return str(value)


def read_expenses(user_folder: str, name: str) -> list[dict]:
    path = _trip_path(user_folder, name)
    if not os.path.exists(path):
        raise TripNotFoundError(name)

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]  # never assume a sheet name (spec §5.1)
    headers = _header_column_map(ws)
    _require_headers(headers, name)

    date_col = headers["Date"]
    amount_col = headers["Amount"]
    category_col = headers["Category"]
    notes_col = headers["Notes"]

    expenses = []
    for row in ws.iter_rows(min_row=2):
        date_val = row[date_col - 1].value
        amount_val = row[amount_col - 1].value
        if date_val is None and amount_val is None:
            continue  # skip fully blank trailing rows
        expenses.append(
            {
                "date": _cell_to_iso_date(date_val),
                "amount": float(amount_val) if amount_val is not None else 0.0,
                "category": str(row[category_col - 1].value or ""),
                "notes": str(row[notes_col - 1].value or ""),
            }
        )
    return expenses


def append_expense(user_folder: str, name: str, expense_date: date, amount: float, category: str, notes: str) -> None:
    path = _trip_path(user_folder, name)
    # Holding the lock across the whole load-modify-save cycle (not just the
    # save) is what actually prevents the race: without it, two threads could
    # both load_workbook() the same pre-append content before either saves.
    with _write_lock:
        if not os.path.exists(path):
            raise TripNotFoundError(name)

        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[0]
        headers = _header_column_map(ws)
        _require_headers(headers, name)

        new_row = ws.max_row + 1
        # Only ever write into the four known columns, wherever they happen to
        # live in this particular file — any other columns on this new row are
        # simply left blank, and every existing row/column is untouched.
        ws.cell(row=new_row, column=headers["Date"], value=expense_date)
        ws.cell(row=new_row, column=headers["Amount"], value=amount)
        ws.cell(row=new_row, column=headers["Category"], value=category)
        ws.cell(row=new_row, column=headers["Notes"], value=notes)
        wb.save(path)
