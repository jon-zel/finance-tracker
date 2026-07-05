"""
One-time migration: translate Hebrew Category values in an existing trip
file to their English equivalents (spec §5.3), so it can be dropped into
./trips/ and work with the English-only UI.

Usage:
    python scripts/migrate_categories.py <input.xlsx> <output.xlsx>

Only the Category column is touched. Date, Amount, Notes (Notes stays in
whatever language it was written in — Hebrew included), formatting, and
any extra columns are preserved exactly.

Extend CATEGORY_MAP below if your files contain category values not
listed here.
"""
import sys

import openpyxl

# Windows terminals often default to a legacy codepage (cp1252) that can't
# print Hebrew category values. Force UTF-8 on stdout/stderr so the warning
# list below never crashes the script on a fresh Windows machine.
for _stream in (sys.stdout, sys.stderr):
    if hasattr(_stream, "reconfigure"):
        _stream.reconfigure(encoding="utf-8", errors="replace")

CATEGORY_MAP = {
    "ארוחות": "Meals",
    "מלון": "Lodging",
    "תחבורה": "Transport",
    "טיסות": "Flights",
    "הוצאות חובה": "Essentials",
    "פנאי": "Leisure",
    "בגדים לעצמי": "Clothing",
    "בגדים": "Clothing",
    "משיכת מזומן": "Cash Withdrawal",
    "מתנות לאחרים": "Gifts for Others",
    "מתנות לעצמי": "Gifts for Self",
    "פאן": "Fun",
    "מאי": "Category MAI",
    "מאי חסכה": "Mai Save",
    "UNKNOWN": "Uncategorized",
}


def find_category_column(ws) -> int:
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip() == "Category":
            return cell.column
    raise ValueError("Could not find a 'Category' column in the input file's header row")


def migrate(input_path: str, output_path: str) -> None:
    wb = openpyxl.load_workbook(input_path)
    ws = wb.worksheets[0]
    category_col = find_category_column(ws)

    unknown_values = set()
    translated_count = 0

    for row in ws.iter_rows(min_row=2):
        cell = row[category_col - 1]
        if cell.value is None:
            continue
        original = str(cell.value).strip()
        if original == "":
            continue
        if original in CATEGORY_MAP:
            cell.value = CATEGORY_MAP[original]
            translated_count += 1
        else:
            unknown_values.add(original)

    wb.save(output_path)

    print(f"Migrated '{input_path}' -> '{output_path}'")
    print(f"  Translated {translated_count} category value(s).")
    if unknown_values:
        print("  WARNING: left the following unknown category value(s) as-is —")
        print("           fix them manually or add them to CATEGORY_MAP:")
        for value in sorted(unknown_values):
            print(f"    - {value!r}")
    else:
        print("  No unknown category values encountered.")


def main() -> None:
    if len(sys.argv) != 3:
        print(f"Usage: python {sys.argv[0]} <input.xlsx> <output.xlsx>")
        sys.exit(1)
    migrate(sys.argv[1], sys.argv[2])


if __name__ == "__main__":
    main()
